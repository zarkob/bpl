#!/usr/bin/env python3
"""
BPL AST to Visio Excel Converter

This script converts a BPL AST JSON file to a Visio-compatible Excel format
for business process visualization.

Usage:
    python ast_to_visio.py input.bpl-ast.json output.xlsx

Dependencies:
    - pandas
    - openpyxl
    - numpy
"""

import json
import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

class BplAstToVisioConverter:
    def __init__(self, ast_json):
        """Initialize with parsed AST json data"""
        self.ast = ast_json
        self.nodes = []
        self.edges = []
        self.swimlanes = []
        
    def normalize_name(self, name):
        """Create a normalized name suitable for Visio"""
        if not name:
            return "unknown"
        return name.lower().replace(" ", "_")
    
    def extract_nodes(self):
        """Extract nodes from AST"""
        for process in self.ast.get('processes', []):
            process_id = process.get('id')
            for lane in process.get('lanes', []):
                lane_id = lane.get('id')
                lane_name = lane.get('name')
                
                # Add lane to swimlanes
                self.swimlanes.append({
                    'id': lane_id,
                    'name': lane_name,
                    'process': process_id
                })
                
                # Process each element in the lane
                for element in lane.get('elements', []):
                    node_type = element.get('type')
                    node_id = element.get('id')
                    node_name = element.get('name')
                    
                    node_data = {
                        'id': node_id,
                        'name': node_name,
                        'type': node_type,
                        'lane': lane_id,
                        'process': process_id
                    }
                    
                    # Add specific attributes based on node type
                    if node_type == 'gateway':
                        node_data['gateway_type'] = element.get('gatewayType', 'exclusive')
                    elif node_type == 'event':
                        node_data['event_type'] = element.get('eventType', 'intermediate')
                    elif node_type == 'task':
                        if 'send:' in node_name:
                            node_data['type'] = 'send'
                        elif 'receive:' in node_name:
                            node_data['type'] = 'receive'
                    
                    self.nodes.append(node_data)
    
    def extract_edges(self):
        """Extract connections from AST"""
        for conn in self.ast.get('connections', []):
            conn_type = conn.get('type')
            source_ref = conn.get('sourceRef')
            target_ref = conn.get('targetRef')
            name = conn.get('name', '')
            
            edge_data = {
                'id': conn.get('id'),
                'type': conn_type,
                'name': name,
                'source': source_ref,
                'target': target_ref
            }
            
            self.edges.append(edge_data)
    
    def process_ast(self):
        """Process the AST and extract all required data"""
        self.extract_nodes()
        self.extract_edges()
        
    def create_visio_dataframes(self):
        """
        Create pandas DataFrames for Visio format
        
        Returns:
            tuple: (nodes_df, edges_df, swimlanes_df)
        """
        # Create DataFrame for nodes
        if self.nodes:
            nodes_df = pd.DataFrame(self.nodes)
            nodes_df['visio_type'] = nodes_df['type'].apply(self.map_node_type_to_visio)
            nodes_df['shape_text'] = nodes_df['name']
            
            # For branch nodes, add label information for connector labels
            branch_nodes = nodes_df[nodes_df['type'] == 'branch'].copy()
            if not branch_nodes.empty and 'label' in branch_nodes.columns:
                # Track branch labels for later use in connection labels
                self.branch_labels = {}
                for _, branch in branch_nodes.iterrows():
                    if 'parentGateway' in branch and 'id' in branch and 'label' in branch:
                        parent = branch['parentGateway']
                        branch_id = branch['id']
                        label = branch['label']
                        self.branch_labels[(parent, branch_id)] = label
            else:
                self.branch_labels = {}
        else:
            nodes_df = pd.DataFrame(columns=['id', 'name', 'type', 'lane', 'process', 'visio_type', 'shape_text'])
            self.branch_labels = {}
        
        # Create DataFrame for edges
        if self.edges:
            edges_df = pd.DataFrame(self.edges)
            edges_df['visio_type'] = edges_df['type'].apply(self.map_edge_type_to_visio)
            
            # Add branch labels to gateway connections
            enhanced_edges = []
            for _, edge in edges_df.iterrows():
                source = edge['source']
                target = edge['target']
                
                # Check if this is a gateway-to-branch connection
                if (source, target) in self.branch_labels:
                    # Use branch label (Yes/No) as the connector label
                    edge_data = edge.copy()
                    edge_data['name'] = self.branch_labels[(source, target)]
                    enhanced_edges.append(edge_data)
                else:
                    enhanced_edges.append(edge)
                    
            # Replace the edges dataframe with enhanced version
            edges_df = pd.DataFrame(enhanced_edges)
            
            # Edges in Visio need source and target shape names
            edges_df['source_shape'] = edges_df['source'].apply(
                lambda src: self.get_node_name_by_id(src, nodes_df))
            edges_df['target_shape'] = edges_df['target'].apply(
                lambda tgt: self.get_node_name_by_id(tgt, nodes_df))
        else:
            edges_df = pd.DataFrame(columns=['id', 'type', 'name', 'source', 'target', 'visio_type', 'source_shape', 'target_shape'])
        
        # Create DataFrame for swimlanes
        if self.swimlanes:
            swimlanes_df = pd.DataFrame(self.swimlanes)
        else:
            swimlanes_df = pd.DataFrame(columns=['id', 'name', 'process'])
        
        return nodes_df, edges_df, swimlanes_df
    
    def get_node_name_by_id(self, node_id, nodes_df):
        """Get node name by ID for edge connections"""
        if nodes_df.empty:
            return ""
        
        matching_nodes = nodes_df[nodes_df['id'] == node_id]
        if not matching_nodes.empty:
            return matching_nodes.iloc[0]['name']
        
        # Handle data objects and other non-standard nodes
        if node_id and node_id.startswith('data_'):
            return node_id.replace('data_', '')
        
        return node_id
    
    def map_node_type_to_visio(self, node_type):
        """Map BPL node types to Visio shape types"""
        type_mapping = {
            'task': 'Process',
            'send': 'Send',
            'receive': 'Receive',
            'gateway': 'Decision',
            'branch': 'Process',
            'event': 'Event',
            'comment': 'Note'
        }
        return type_mapping.get(node_type, 'Process')
    
    def map_edge_type_to_visio(self, edge_type):
        """Map BPL edge types to Visio connector types"""
        type_mapping = {
            'sequenceFlow': 'Sequence',
            'messageFlow': 'Message',
            'dataAssociation': 'Data'
        }
        return type_mapping.get(edge_type, 'Sequence')
    
    def generate_excel(self, output_path):
        """Generate Visio-compatible Excel file"""
        nodes_df, edges_df, swimlanes_df = self.create_visio_dataframes()
        
        # Create a workbook directly with openpyxl instead of using pandas ExcelWriter
        wb = Workbook()
        ws = wb.active
        ws.title = "Process Diagram"
        
        # Use customer-friendly format for business process flows matching the requested format
        # Column headers
        headers = [
            "Process Step ID", 
            "Process Step Description", 
            "Next Step ID", 
            "Connector Label", 
            "Shape Type", 
            "Function", 
            "Phase", 
            "Owner", 
            "Cost", 
            "Start Date", 
            "End Date", 
            "Status", 
            "Alt Description"
        ]
        
        # Add the headers to the first row
        ws.append(headers)
        
        # Format headers with bold font
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # Build connection lookup for efficient node connection mapping
        connection_map = {}
        for _, edge in edges_df.iterrows():
            source_id = edge['source']
            target_id = edge['target']
            connection_label = edge['name']
            
            if source_id not in connection_map:
                connection_map[source_id] = []
                
            connection_map[source_id].append({
                'target': target_id,
                'label': connection_label
            })
            
        # Add process steps (nodes) with their connections
        for _, node in nodes_df.iterrows():
            node_id = node['id']
            
            # Get all targets for this node
            targets = []
            connector_labels = []
            if node_id in connection_map:
                for conn in connection_map[node_id]:
                    target_id = conn['target']
                    
                    # Clean up target ID similar to source ID
                    clean_target = target_id
                    if '_' in target_id:
                        target_parts = target_id.split('_', 1)
                        if len(target_parts) > 1:
                            clean_target = target_parts[1]
                    
                    # Add uppercase version of clean target ID
                    targets.append(clean_target.upper())
                    
                    # Always capture connection labels, prioritizing branch annotations (Yes/No)
                    label = conn['label'] if conn['label'] else ""
                    
                    # If this is a gateway branch connection, check for branch labels
                    source_id = node_id
                    target_id = conn['target']
                    
                    # First check if this connection has a branch label like Yes/No
                    if hasattr(self, 'branch_labels') and (source_id, target_id) in self.branch_labels:
                        label = self.branch_labels[(source_id, target_id)]
                    
                    # For gateway-branch connections, use branch label (Yes/No)
                    elif label.lower() == "by fields" or not label:
                        # Try to identify gateway connections from the node types
                        source_node = next((n for n in self.nodes if n['id'] == source_id), None)
                        target_node = next((n for n in self.nodes if n['id'] == target_id), None)
                        
                        if source_node and target_node:
                            if source_node.get('type') == 'gateway' and target_node.get('type') == 'branch':
                                # This is a gateway-branch connection, check if target has a label
                                if 'label' in target_node:
                                    label = target_node['label']
                    
                    # Add the label (empty string if none found)
                    connector_labels.append(label)
                        
            next_step_id = ",".join(targets) if targets else ""
            
            # Include all meaningful connector labels, filtering out "by Fields" and empty labels
            filtered_labels = []
            for label in connector_labels:
                if label and label.lower() != "by fields":
                    filtered_labels.append(label)
                    
            connector_label = ",".join(filtered_labels)
            
            # Map BPL node type to Visio shape type
            shape_type = "Process"  # Default shape type
            
            # Get the node type from the dataframe
            node_type = node['type'] if 'type' in node else ""
            
            # Map node types to shape types
            if node_type == 'gateway':
                shape_type = "Decision"
            elif node_type == 'event':
                event_type = node.get('eventType', '')
                if event_type == 'start':
                    shape_type = "Start"
                elif event_type == 'end':
                    shape_type = "End"
                else:
                    shape_type = "Process"  # Default for intermediate events
            elif node_type == 'comment':
                shape_type = "Document"
            elif node_type == 'dataObject':
                shape_type = "Data"
            elif node_type == 'branch':
                # Branches are not separate shapes in Visio - they're represented as connections
                shape_type = "Custom 1"
            elif node_type == 'send' or node_type == 'receive':
                shape_type = "External reference"  # For message handling
            
            # Function corresponds to Lane in BPL
            function = node['lane'] if 'lane' in node else ""
            
            # Format the description as "ID : Name" per the requested format 
            # Remove any lane prefixes from the node_id for cleaner IDs
            clean_id = node_id
            if '_' in node_id and function:
                parts = node_id.split('_', 1)
                if parts[0].lower() == function.lower():
                    clean_id = parts[1]
                    
            # Create standardized step ID - prefer to use just the task name without lane prefix
            step_id = clean_id.upper()
            
            # Format description as "STEP_ID : Step name" per the requested format
            description = f"{step_id} : {node['name']}" if 'name' in node else step_id
            
            # Create a row with all the columns
            ws.append([
                step_id,                  # Process Step ID
                description,              # Process Step Description
                next_step_id,             # Next Step ID
                connector_label,          # Connector Label
                shape_type,               # Shape Type
                function,                 # Function (Lane)
                "",                       # Phase
                "",                       # Owner
                "",                       # Cost
                "",                       # Start Date
                "",                       # End Date
                "",                       # Status
                ""                        # Alt Description
            ])
        
        # Adjust column widths for better visibility
        col_widths = {
            'A': 15,  # Process Step ID
            'B': 40,  # Process Step Description
            'C': 40,  # Next Step ID
            'D': 20,  # Connector Label
            'E': 15,  # Shape Type
            'F': 25,  # Function
            'G': 15,  # Phase
            'H': 15,  # Owner
            'I': 10,  # Cost
            'J': 15,  # Start Date
            'K': 15,  # End Date
            'L': 15,  # Status
            'M': 25,  # Alt Description
        }
        
        # Set column widths
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
            
        # Define a named range for Visio import
        # Get the number of rows with data (header + data rows)
        num_rows = len(list(ws.rows))
        # Get the number of columns (13 columns in the defined format)
        num_cols = 13
        
        # Create a named range covering all data cells
        # Range format is worksheetname!$A$1:$M$last_row
        from openpyxl.workbook.defined_name import DefinedName
        
        # Define the range reference string
        range_reference = f"'{ws.title}'!$A$1:${chr(64 + num_cols)}${num_rows}"
        
        # Create a named range for Visio
        defined_name = DefinedName('Visio_01', attr_text=range_reference)
        
        # Add the named range to the workbook
        wb.defined_names.add(defined_name)
        
        # Save the workbook
        wb.save(output_path)
        
        print(f"Excel file generated at: {output_path}")
        print(f"Created named range 'Visio_01' covering {range_reference}")
        return output_path

def main():
    if len(sys.argv) < 3:
        print("Usage: python ast_to_visio.py input.bpl-ast.json output.xlsx")
        sys.exit(1)
    
    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    if not os.path.exists(input_json):
        print(f"Error: Input file {input_json} not found.")
        sys.exit(1)
    
    try:
        with open(input_json, 'r') as f:
            ast_data = json.load(f)
        
        converter = BplAstToVisioConverter(ast_data)
        converter.process_ast()
        converter.generate_excel(output_xlsx)
        
        print(f"Successfully converted {input_json} to Visio format at {output_xlsx}")
    except Exception as e:
        print(f"Error processing the AST: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()